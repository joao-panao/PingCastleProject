import openpyxl
from bs4 import BeautifulSoup

# Load HTML file
with open("C:/Users/joao.panao/OneDrive - RIS 2048/Documentos/GitHub/PingCastleProject/Ping Castle Reports/ad_hc_cmas.local.html", encoding="utf8") as fp:
    soup = BeautifulSoup(fp, "html.parser")

# Find div with id="sectionPrivilegedAccounts"
priv_section = soup.find("div", {"id": "sectionPrivilegedAccounts"})

# Find all button tags and extract their text
priv_data = []
for button in priv_section.find_all("button"):
    priv_data.append(button.text)

# Find div with id="sectionAnomaliesanalysis"
anom_section = soup.find("div", {"id": "sectionAnomaliesanalysis"})

# Find all button tags and extract their text
anom_data = []
for button in anom_section.find_all("button"):
    anom_data.append(button.text)

# Find div with id="usersaccordion"
usersaccordion_div = soup.find("div", {"id": "usersaccordion"})

# Find all button tags and extract their text
usersaccordion_data = []
for button in usersaccordion_div.find_all("button"):
    usersaccordion_data.append(button.text)

# Find all i tags and extract their text
i_data = []
for i in usersaccordion_div.find_all("i"):
    i_data.append(i.text[1:-1])

# Find the div with ID="rulesStaleObjects"
div = soup.find("div", {"id": "rulesStaleObjects"})

# Get the text of the buttons inside the div
button_texts = [button.get_text(strip=True) for button in div.find_all("button")]

# Write data to Excel file
wb = openpyxl.Workbook()

# Write privileged accounts data to sheet 1
ws1 = wb.active
ws1.title = "Privileged Accounts"
for i, val in enumerate(priv_data, start=1):
    ws1.cell(row=i, column=1, value=val)

# Write anomalies analysis data to sheet 2
ws2 = wb.create_sheet(title="Anomalies Analysis")
for i, val in enumerate(anom_data, start=1):
    ws2.cell(row=i, column=1, value=val)

# Write users accordion data and i tag data to sheet 3
ws3 = wb.create_sheet(title="Users Accordion")
for i, val in enumerate(usersaccordion_data, start=1):
    ws3.cell(row=i, column=1, value=val)
for i, val in enumerate(i_data, start=1):
    ws3.cell(row=i, column=2, value=val)

# Open the existing workbook and create a new worksheet
ws4 = wb.create_sheet(title="Stale Objects")
for i, text in enumerate(button_texts):
    ws4.cell(row=i+1, column=1, value=text)

wb.save("C:/Users/joao.panao/OneDrive - RIS 2048/Documentos/GitHub/PingCastleProject/Extracted Tables/Combined Data4.xlsx")