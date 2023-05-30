import openpyxl
from bs4 import BeautifulSoup

# Load HTML file
with open("C:/Users/joao.panao/OneDrive - RIS 2048/Documentos/GitHub/PingCastleProject/Ping Castle Reports/ad_hc_cmas.local.html", encoding="utf8") as fp:
    soup = BeautifulSoup(fp, "html.parser")

# Find div with id="sectionPrivilegedAccounts"
priv_section = soup.find("div", {"id": "sectionPrivilegedAccounts"})

# Find all button tags and extract their text
priv_data = []
for button in priv_section.find_all("button"):
    priv_data.append(button.text)

# Find div with id="sectionAnomaliesanalysis"
anom_section = soup.find("div", {"id": "sectionAnomaliesanalysis"})

# Find all button tags and extract their text
anom_data = []
for button in anom_section.find_all("button"):
    anom_data.append(button.text)

# Write data to Excel file
wb = openpyxl.Workbook()

# Write privileged accounts data to sheet 1
ws1 = wb.active
ws1.title = "Privileged Accounts"
for i, val in enumerate(priv_data, start=1):
    ws1.cell(row=i, column=1, value=val)

# Write anomalies analysis data to sheet 2
ws2 = wb.create_sheet(title="Anomalies Analysis")
for i, val in enumerate(anom_data, start=1):
    ws2.cell(row=i, column=1, value=val)

wb.save("C:/Users/joao.panao/OneDrive - RIS 2048/Documentos/GitHub/PingCastleProject/Extracted Tables/Combined Data.xlsx")
